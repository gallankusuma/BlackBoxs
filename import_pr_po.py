#!/usr/bin/env python3
"""
Parse 2. PR 2026.xlsx and 4. PO 2026.xlsx → generate SQL for VPS insert
"""
import openpyxl, re
from datetime import datetime, date

REQUESTOR_ID = 1   # admin
STATUS_PR    = 'APPROVED'
STATUS_PO    = 'submitted'

def clean(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, date): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s if s else None

def esc(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'","''").replace("\\","\\\\") + "'"

def num(v, default=0):
    try: return float(v)
    except: return default

# ─── Parse PR Workbook ────────────────────────────────────────────────────────
wb_pr = openpyxl.load_workbook('2. PR 2026.xlsx', data_only=True)

prs = []
for sh in wb_pr.sheetnames:
    ws = wb_pr[sh]
    # Find PR number from header area
    pr_no = None; pr_date = None; needed_by = None; notes = None
    for r in range(1, 10):
        for c in range(1, 13):
            val = ws.cell(r, c).value
            if isinstance(val, str) and 'No PR:' in val:
                m = re.search(r'No PR:\s*([\w\-]+)', val)
                if m: pr_no = 'GEN-PR/' + m.group(1)
        # dates
        row_vals = [ws.cell(r, c).value for c in range(1, 13)]
        row_str  = ' '.join(str(v) for v in row_vals if v)
        if 'Tanggal PR' in row_str:
            for c in range(1, 13):
                v = ws.cell(r, c).value
                if isinstance(v, datetime): pr_date = v.strftime('%Y-%m-%d'); break
        if 'Tanggal Delivery' in row_str or 'Delivery' in row_str:
            for c in range(1, 13):
                v = ws.cell(r, c).value
                if isinstance(v, datetime): needed_by = v.strftime('%Y-%m-%d'); break
        if 'Jenis PR' in row_str:
            for c in range(1, 13):
                v = ws.cell(r, c).value
                if isinstance(v, str) and len(v) > 3 and 'Jenis' not in v and ':' not in v:
                    notes = v.strip(); break

    if not pr_no:
        pr_no = f'GEN-PR/{sh}'

    # Find item rows (numbered items)
    items = []
    for r in range(8, ws.max_row + 1):
        no_val   = ws.cell(r, 1).value
        name_val = ws.cell(r, 2).value
        # item row: col1 is int/digit-str, col2 is non-empty item name
        if no_val is None or name_val is None: continue
        no_str = str(no_val).strip()
        if not re.match(r'^\d+$', no_str): continue
        name_str = str(name_val).strip()
        if len(name_str) < 2: continue
        if any(x in name_str.upper() for x in ['DIBUAT','APPROVE','WAKTU']): continue

        # qty: col 6 or nearby
        qty_raw  = ws.cell(r, 6).value
        uom_raw  = ws.cell(r, 7).value
        spec_raw = ws.cell(r, 5).value
        price_raw= ws.cell(r, 8).value

        # Parse qty (may be "10 sak" format)
        qty = 1; uom = 'pcs'
        if qty_raw is not None:
            m = re.match(r'(\d+(?:\.\d+)?)\s*(\w*)', str(qty_raw).strip())
            if m:
                qty = float(m.group(1)) if m.group(1) else 1
                if m.group(2): uom = m.group(2)
        if uom_raw and str(uom_raw).strip():
            uom = str(uom_raw).strip()

        spec = clean(spec_raw) or ''
        full_name = name_str + (f' ({spec})' if spec else '')
        unit_price = num(price_raw, 0)

        items.append({
            'name': full_name[:200],
            'qty':  qty,
            'uom':  uom[:20] if uom else 'pcs',
            'unit_price': unit_price,
        })

    prs.append({
        'pr_no':      pr_no,
        'pr_date':    pr_date or '2026-05-01',
        'needed_by':  needed_by or pr_date or '2026-05-07',
        'notes':      notes or sh,
        'items':      items,
    })

# ─── Parse PO Workbook ────────────────────────────────────────────────────────
wb_po = openpyxl.load_workbook('4. PO 2026.xlsx', data_only=True)

pos = []
for sh in wb_po.sheetnames:
    ws = wb_po[sh]
    # Row 6: To: <vendor>, Row 7: PO No, Row 8: Date
    vendor_name = clean(ws.cell(6, 3).value) or 'UNKNOWN'
    po_no_raw   = ws.cell(7, 13).value
    po_date_raw = ws.cell(8, 13).value
    total_raw   = None

    po_no   = clean(po_no_raw) or f'GEN-PO/{sh}'
    po_date = clean(po_date_raw) or '2026-05-01'

    # Find item rows starting around row 15-17
    items = []
    for r in range(14, ws.max_row + 1):
        no_val   = ws.cell(r, 2).value
        qty_val  = ws.cell(r, 3).value
        uom_val  = ws.cell(r, 4).value
        name_val = ws.cell(r, 5).value
        spec_val = ws.cell(r, 11).value
        price_val= ws.cell(r, 12).value
        total_val= ws.cell(r, 14).value

        if no_val is None or name_val is None: continue
        no_str = str(no_val).strip()
        if not re.match(r'^\d+$', no_str): continue
        name_str = str(name_val).strip()
        if len(name_str) < 2: continue
        if any(x in name_str.upper() for x in ['AMOUNT','TOTAL','PPN','GRAND']): continue

        spec = clean(spec_val) or ''
        full_name = name_str + (f' ({spec})' if spec else '')
        qty       = num(qty_val, 1)
        uom       = str(uom_val).strip() if uom_val else 'pcs'
        unit_price= num(price_val, 0)
        line_total= num(total_val, qty * unit_price)

        items.append({
            'name':       full_name[:200],
            'qty':        qty,
            'uom':        uom[:20],
            'unit_price': unit_price,
            'line_total':  line_total,
        })

        # Check for TOTAL row
        if 'TOTAL' in str(name_val).upper() and not re.match(r'^\d+$', no_str):
            break

    # Grand total from row ~30-32
    for r in range(25, min(ws.max_row+1, 40)):
        for c in [12,13,14]:
            v = ws.cell(r, c).value
            if v and str(v).strip() in ['GRAND TOTAL ','GRAND TOTAL']:
                total_raw = ws.cell(r, 14).value or ws.cell(r, 13).value
    total_amount = num(total_raw, sum(i['line_total'] for i in items))

    pos.append({
        'po_no':        po_no,
        'po_date':      po_date,
        'vendor_name':  vendor_name.strip(),
        'total_amount': total_amount,
        'items':        items,
        'payment_term': '100% BEFORE SHIPMENT',
        'ppn_percent':  0,
    })

# ─── Generate SQL ─────────────────────────────────────────────────────────────
lines = []
lines.append("SET FOREIGN_KEY_CHECKS=0;")
lines.append("SET SQL_MODE='NO_AUTO_VALUE_ON_ZERO';")

# Collect unique vendors from PO
all_vendors = sorted(set(p['vendor_name'] for p in pos))
lines.append("\n-- ============================================================")
lines.append("-- 1. VENDORS")
lines.append("-- ============================================================")
for i, vname in enumerate(all_vendors):
    code = f'GEN-V{i+1:03d}'
    lines.append(f"INSERT IGNORE INTO vendors (code, name, supply_category, payment_terms, is_active) VALUES ({esc(code)}, {esc(vname)}, 'Material', '100% Before Shipment', 1);")

# Collect unique items → create products
all_items = {}
for pr in prs:
    for it in pr['items']:
        key = it['name'].upper()[:100]
        if key not in all_items:
            all_items[key] = it['name']
for po in pos:
    for it in po['items']:
        key = it['name'].upper()[:100]
        if key not in all_items:
            all_items[key] = it['name']

lines.append("\n-- ============================================================")
lines.append("-- 2. PRODUCTS (items not yet in master)")
lines.append("-- ============================================================")
for i, (key, name) in enumerate(all_items.items()):
    sku = f'GEN-MAT-{i+1:04d}'
    lines.append(f"INSERT IGNORE INTO products (sku, name, product_type_id, is_active) SELECT {esc(sku)}, {esc(name)}, 1, 1 WHERE NOT EXISTS (SELECT 1 FROM products WHERE name = {esc(name)});")

lines.append("\n-- ============================================================")
lines.append("-- 3. PURCHASE REQUESTS")
lines.append("-- ============================================================")
for pr in prs:
    lines.append(f"""
-- PR: {pr['pr_no']}
INSERT IGNORE INTO purchase_requests (pr_number, requestor_id, status, approval_status, request_date, needed_by, notes, approval_required)
VALUES ({esc(pr['pr_no'])}, {REQUESTOR_ID}, {esc(STATUS_PR)}, 2, {esc(pr['pr_date'])}, {esc(pr['needed_by'])}, {esc(pr['notes'])}, 0);
SET @pr_id = (SELECT id FROM purchase_requests WHERE pr_number = {esc(pr['pr_no'])});""")
    for it in pr['items']:
        lines.append(f"""INSERT INTO purchase_request_items (purchase_request_id, product_id, quantity, unit_price, notes)
SELECT @pr_id, p.id, {it['qty']}, {it['unit_price']}, {esc(it['uom'])}
FROM products p WHERE p.name = {esc(it['name'])} LIMIT 1;""")

lines.append("\n-- ============================================================")
lines.append("-- 4. PURCHASE ORDERS")
lines.append("-- ============================================================")
for po in pos:
    lines.append(f"""
-- PO: {po['po_no']} | Vendor: {po['vendor_name']}
SET @v_id = (SELECT id FROM vendors WHERE name = {esc(po['vendor_name'])} LIMIT 1);
INSERT IGNORE INTO purchase_orders (po_number, vendor_id, po_date, status, total_amount, payment_term, ppn_percent, approval_required, approval_status, notes)
VALUES ({esc(po['po_no'])}, @v_id, {esc(po['po_date'])}, {esc(STATUS_PO)}, {po['total_amount']}, {esc(po['payment_term'])}, {po['ppn_percent']}, 0, 2, NULL);
SET @po_id = (SELECT id FROM purchase_orders WHERE po_number = {esc(po['po_no'])});""")
    for it in po['items']:
        lines.append(f"""INSERT INTO purchase_order_items (purchase_order_id, product_id, quantity, unit_price, line_total, uom)
SELECT @po_id, p.id, {it['qty']}, {it['unit_price']}, {it['line_total']}, {esc(it['uom'])}
FROM products p WHERE p.name = {esc(it['name'])} LIMIT 1;""")

lines.append("\nSET FOREIGN_KEY_CHECKS=1;")
lines.append("SELECT 'IMPORT COMPLETE' AS status;")

sql = '\n'.join(lines)
with open('/tmp/import_pr_po.sql', 'w') as f:
    f.write(sql)

print(f"Generated SQL: {len(lines)} statements")
print(f"  PRs: {len(prs)}, POs: {len(pos)}")
print(f"  Vendors: {len(all_vendors)}, Items: {len(all_items)}")
print("  → /tmp/import_pr_po.sql")
